from functools import partial
from itertools import zip_longest
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException, WebDriverException
import getpass
import sys
import easygui
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from decouple import config




class Dados:

    def tratamento_de_usuario(self):
        op = input('Tentar novamente (s/n)?: ')
        if op in ['s','n','S','N']:
            if op == 's' or op == 'S':
                print()
                usuario, senha = self.dados_usuario()
                return usuario , senha
            if op == 'n' or op == 'N':
                print('Programa Finalizado!')
                self.driver.quit()
                sys.exit()
        else:
            print('Opção invalida:')
            print('Programa Finalizado!')
            self.driver.quit()
            sys.exit()
            
    def dados_usuario(self):
        
        login = input('Insira email de acesso: ')
        password = input('Insira senha de acesso: ')

        if login == '' or password == '':
            print('Preencha todos os campos corretamente!!')
            self.tratamento_de_usuario()
        else:    
            return login, password


    def valida_login(self,usuario, senha):
        if usuario == '' or  senha == '':
            print('Existem campos Vazios!!')
            self.valida_login(usuario,senha)
        else:
            return usuario, senha
   
    def login_asap(self):
        url='https://app.asaplog.com.br/'
        self.driver = webdriver.Chrome(ChromeDriverManager().install())

        self.driver.get(url)
        op = True
        while op:
            user = self.driver.find_element_by_id('username')
            password = self.driver.find_element_by_id('password')
            btn_login = self.driver.find_element_by_id('submit')
            logado = self.driver.find_element_by_id('remember_me')
            self.usuario, self.senha = self.dados_usuario()
            user.send_keys(self.usuario)
            password.send_keys(self.senha)
            logado.click()
            btn_login.click()
            erro = self.driver.find_elements_by_xpath('//*[@id="login"]/div/div[2]/div[1 and (@class ="alert alert-danger")]')
            if len(erro)>0:
                print(erro[0].text)
                print('')
                self.usuario, self.senha = self.tratamento_de_usuario()
                if self.usuario == '' and self.senha == '':
                    op = False
                    print('Programa Finalizado')
                else:
                    op = True
            else:
                self.roteiro_direto(self.driver)
                self.extrair_dados_tabela()
                self.driver.quit()
                sys.exit()
      
    def roteiro_direto(self,driver):
        menu = driver.find_element_by_xpath('//*[@id="menu-monitorar"]/a')
        roteiro_direto = driver.find_element_by_xpath('//*[@id="menu-monitorar"]/ul/li[3]/a')
        menu.click()
        roteiro_direto.click()

    def qtde_paginas(self):
        sleep(5)
        pagina = self.driver.find_element_by_id('RoteirosDia_info')
        pagina = str(pagina.text)
        num_pagina = int(pagina.split()[4])
        if num_pagina == 0:
            print('não existe conteudo')
            self.driver.quit()
        else:
            num_pagina
            return  num_pagina

    def qtde_linhas(self):
        sleep(2)
        linhas = self.driver.find_elements_by_xpath('//table/tbody/tr')
        num_lin = len(linhas)
        return num_lin

    def extrair_dados_tabela(self):
        wdw = WebDriverWait(self.driver,15)
        TRANSPORTADORA = []
        CODIGO = []
        HUB = []
        PARADAS = []
        PEDIDOS = []
        DISTANCIA = []
        CUBAGEM = []
        ENTREGADOR = []
        COLETAS_ENTREGAS = []

        pg = self.qtde_paginas()

        for pagina in range(1,(pg + 1)):

            lin = self.qtde_linhas()
            print(f'extraindo dados: {pagina} de {pg} pag | linhas {lin} ...')
            for l in range(1,(lin + 1)):
                for c in range(1,10):
                    valor_campo = '//table/tbody/tr['+ str(l) + ']/td['+ str(c) + ']'

                    try:
                        element = self.driver.find_element_by_xpath(valor_campo)
                    except (StaleElementReferenceException, WebDriverException) as e:
                        espera = partial(self.esperar_elemento,By.XPATH, '//*[@id="hojeDireto"]/div/div[1]/div[1]/button')
                        valida = wdw.until(espera)
                        if valida == True:
                            element = self.driver.find_element_by_xpath(valor_campo)
                            continue
                        else:
                            print(f'erro na pagina {pg} linha {l}')
                            sys.exit()



                    if c == 1:
                        CODIGO.append(element.text)
                    elif c == 2:
                        TRANSPORTADORA.append(element.text)
                    elif c == 3:
                        HUB.append(element.text) 
                    elif c == 4:
                        PARADAS.append(element.text)
                    elif c == 5:
                        PEDIDOS.append(element.text)
                    elif c == 6:
                        DISTANCIA.append(element.text)
                    elif c == 7:
                        CUBAGEM.append(element.text)
                    elif c == 8:
                        ENTREGADOR.append(element.text) 
                    elif c == 9:
                        COLETAS_ENTREGAS.append(element.text)



                    else:
                        continue
            sleep(2)
            espera_btn = partial(self.esperar_elemento,By.XPATH, '//*[@id="RoteirosDia_next"]')
            valida_btn = wdw.until(espera_btn)
            if valida_btn == True:
                btn_prox = self.driver.find_element_by_xpath('//*[@id="RoteirosDia_next"]')
                btn_prox.click()
            else:
                print('Não foi possivel localiza a proxima pagina!')
                self.driver.quit()
                sys.exit()

        op2 = input('Gerar arquivo Excel (s/n)?: ')
        if op2 in ['s','n','S','N']:
            if op2 == 's' or op2 == 'S':
                print()
                diretorio = self.caminho_arquivo_excel()
                if diretorio == '':
                    print('Caminho definido não é valido... Dados não salvos!')
                    self.driver.quit()
                    sys.exit()

                else:
                    self.cria_planilha(diretorio,CODIGO,TRANSPORTADORA, HUB,PARADAS,PEDIDOS,DISTANCIA,CUBAGEM,ENTREGADOR, COLETAS_ENTREGAS)
                    print('Processo Concluido')
            elif op2 == 'n' or op2 == 'N':
                print('Dados não salvos!')
                self.driver.quit()
                sys.exit()
            else:
                print('Dados não salvos!')
                self.driver.quit()
                sys.exit()
    

    def cria_planilha(self,dirSave,CODIGO,TRANSPORTADORA, HUB,PARADAS,PEDIDOS,DISTANCIA,CUBAGEM,ENTREGADOR, COLETAS_ENTREGAS):
        index = 2
        wb = openpyxl.Workbook()
        active_sheet = wb.active  # Pegando sheet ativa
        active_sheet.title = "BASE_DADOS"  # Mudando titulo
        active_sheet['A1'] = 'CODIGO'
        active_sheet['B1']='TRANSPORTADORA'
        active_sheet['C1']='HUB'
        active_sheet['D1']='PARADAS'
        active_sheet['E1']='PEDIDOS'
        active_sheet['F1']='DISTÂNCIA'
        active_sheet['G1']='CUBAGEM'
        active_sheet['H1']='ENTREGADOR'
        active_sheet['I1']='COLETAS_ENTREGAS'

        resumo_tabela = zip_longest(CODIGO, TRANSPORTADORA, HUB, PARADAS, PEDIDOS, DISTANCIA, CUBAGEM, ENTREGADOR, COLETAS_ENTREGAS)
        for codigo, transportadora, hub, parada, pedidos, distancia, cubagem, entregador, coleta_entre in  resumo_tabela:

            active_sheet.cell(column=1, row=index, value=codigo)
            active_sheet.cell(column=2,row=index,value=transportadora)
            active_sheet.cell(column=3,row=index,value=hub)
            active_sheet.cell(column=4,row=index,value=parada)
            active_sheet.cell(column=5,row=index,value=pedidos)
            active_sheet.cell(column=6,row=index,value=distancia)
            active_sheet.cell(column=7,row=index,value=cubagem)
            active_sheet.cell(column=8,row=index,value=entregador)
            active_sheet.cell(column=9,row=index,value=coleta_entre)

            index += 1
        data_hora = datetime.now()
        data_format = data_hora.strftime('%d%m%Y_%H%M%S')
        wb.save(dirSave + 'Roteiro_direto_' + data_format +'.xlsx')

    def esperar_elemento(by, element, driver):
        return bool(driver.find_elements(by,element))

    def caminho_arquivo_excel(self):
        diretorio = easygui.diropenbox()
        return diretorio +'\\'

    def esperar_elemento(self, by, element, driver):
        return bool(driver.find_elements(by,element))



if __name__ == '__main__':
    extrair_dados = Dados()
    extrair_dados.login_asap()


